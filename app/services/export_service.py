from pathlib import Path
from shutil import copy2

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from app.config import LEGACY_WORKBOOK_PATH, OUTPUT_DIR

try:
    import win32com.client  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    win32com = None


def export_excel(state):
    path = OUTPUT_DIR / "resultado_convalidacion.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Cabecera"
    header = state.header
    rows = [
        ("Formato", header.formato),
        ("Código estudiante", header.codigo_estudiante),
        ("Apellidos", header.apellidos),
        ("Nombres", header.nombres),
        ("Carrera código", header.carrera_codigo),
        ("Carrera nombre", header.carrera_nombre),
        ("Modalidad", header.modalidad),
        ("Sede", header.sede),
        ("Año validez", header.anio_validez),
        ("Centro procedencia", header.origen_nombre_procedencia),
        ("Carrera procedencia", header.origen_carrera_procedencia),
        ("Año validez origen", header.origen_anio_validez),
        ("Nota aprobatoria origen", header.origen_nota_aprobatoria),
        ("Fecha", header.fecha),
        ("Cargo revisor", header.cargo_revisor),
        ("Nivel académico", header.nivel_academico),
        ("Observaciones", header.observaciones),
    ]
    for r in rows:
        ws1.append(r)

    ws2 = wb.create_sheet("Cursos Origen")
    ws2.append(["Código", "Nombre", "Créditos", "Nota", "Aprobado", "Tipo"])
    for c in state.source_courses:
        ws2.append([c.codigo, c.nombre, c.creditos, c.nota, "Sí" if c.aprobado else "No", c.tipo])

    ws3 = wb.create_sheet("Mapeos")
    ws3.append(["Origen código", "Origen nombre", "Destino código", "Destino nombre", "Cred. origen", "Cred. destino", "Score", "Estado", "Observación"])
    for m in state.mappings:
        ws3.append([m.origen_codigo, m.origen_nombre, m.destino_codigo, m.destino_nombre, m.creditos_origen, m.creditos_destino, m.score, m.estado, m.observacion])

    if state.selected_malla:
        ws4 = wb.create_sheet("Malla Destino")
        ws4.append(["Ciclo", "Código", "Nombre", "Créditos", "Horas", "Requisitos"])
        for c in state.selected_malla.get("courses", []):
            ws4.append([c.get("cycle"), c.get("code"), c.get("name"), c.get("credits"), c.get("hours"), c.get("requirements")])

    wb.save(path)
    return path


def _build_dest_lookup(state):
    lookup = {}
    if not state.selected_malla:
        return lookup

    for course in state.selected_malla.get("courses", []):
        code = str(course.get("code") or "").strip()
        if code:
            lookup[code] = course
    return lookup


def _legacy_header_values(state):
    h = state.header
    return {
        "GralFormato": h.formato,
        "GralCodEstudiante": h.codigo_estudiante,
        "GralApellidosEstudiante": h.apellidos,
        "GralNombresEstudiante": h.nombres,
        "GralFecha": h.fecha,
        "GralOrigenNotaAprobatoria": h.origen_nota_aprobatoria,
        "GralObservacionConva": h.observaciones,
        "GralSede": h.sede,
        "GralCarrera": h.carrera_nombre,
        "GralModalidad": h.modalidad,
        "GralAnyoValidez": h.anio_validez,
        "GralCodCarrera": h.carrera_codigo,
        "GralOrigenNombreProcedencia": h.origen_nombre_procedencia,
        "GralOrigenCarreraProcedencia": h.origen_carrera_procedencia,
        "GralOrigenAnyoValidez": h.origen_anio_validez,
        "GralCargoRevisor": h.cargo_revisor,
    }

def _fill_legacy_detail_sheet_excel(workbook, state):
    ws = workbook.Worksheets("Detalle de Equivalencias")
    source_lookup = {item.codigo: item for item in state.source_courses}
    dest_lookup = _build_dest_lookup(state)

    ws.Range("M9:AE120").ClearContents()

    for idx, mapping in enumerate(state.mappings[:30], start=1):
        row = 8 + idx
        source = source_lookup.get(mapping.origen_codigo)
        dest = dest_lookup.get(mapping.destino_codigo, {})

        ws.Range(f"M{row}").Value = idx
        ws.Range(f"N{row}").Value = idx
        ws.Range(f"P{row}").Value = mapping.destino_codigo
        ws.Range(f"Q{row}").Value = mapping.origen_codigo
        ws.Range(f"R{row}").Value = mapping.origen_nombre
        ws.Range(f"S{row}").Value = mapping.creditos_origen
        ws.Range(f"T{row}").Value = getattr(source, "nota", 0)
        ws.Range(f"Y{row}").Value = mapping.destino_codigo
        ws.Range(f"AA{row}").Value = mapping.destino_nombre
        ws.Range(f"AB{row}").Value = mapping.creditos_destino
        ws.Range(f"AC{row}").Value = mapping.destino_codigo[:4]
        ws.Range(f"AD{row}").Value = mapping.destino_codigo
        ws.Range(f"AE{row}").Value = dest.get("cycle", "")


def _legacy_sheet_names(formato):
    formato = (formato or "").upper()
    if formato == "CONVALIDACION Y RECOMENDACION":
        return ["Convalidacion", "Recomendacion"]
    if formato == "AMPLIACION DE CONVALIDACION Y RECOMENDACION":
        return ["Reconvalidacion", "Recomendacion"]
    if formato == "REINGRESO SOLAMENTE":
        return ["Reingreso"]
    if formato == "REINGRESO Y CAMBIO CARRERA/MODALIDAD/SEDE":
        return ["Reingreso", "CambioCarreraModalidadSede"]
    return ["Convalidacion"]


def _export_pdf_with_excel_template(state):
    if not LEGACY_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla legado: {LEGACY_WORKBOOK_PATH}")
    if win32com is None:
        raise RuntimeError("win32com no está disponible para exportación legado.")

    template_copy = OUTPUT_DIR / "legacy_export_tmp.xlsm"
    copy2(LEGACY_WORKBOOK_PATH, template_copy)

    excel = None
    wb = None
    generated_paths = []
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(template_copy.resolve()))
        for name, value in _legacy_header_values(state).items():
            try:
                wb.Names.Item(name).RefersToRange.Value = value
            except Exception:
                pass

        _fill_legacy_detail_sheet_excel(wb, state)
        wb.RefreshAll()
        excel.CalculateFull()

        for sheet_name in _legacy_sheet_names(state.header.formato):
            output_path = OUTPUT_DIR / f"{sheet_name.lower()}_{state.header.codigo_estudiante or 'sin_codigo'}.pdf"
            ws = wb.Worksheets(sheet_name)
            ws.ExportAsFixedFormat(0, str(output_path.resolve()), 0, True, True, None, None, False)
            generated_paths.append(output_path)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()

    if not generated_paths:
        raise RuntimeError("No se generó ningún PDF legado.")

    if len(generated_paths) == 1:
        return generated_paths[0]
    return " | ".join(str(path) for path in generated_paths)


def _export_pdf_simple(state):
    path = OUTPUT_DIR / "resultado_convalidacion.pdf"
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4

    y = height - 18 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, "Resumen de Convalidación / Reingreso")
    y -= 8 * mm

    c.setFont("Helvetica", 9)
    header_pairs = [
        ("Formato", state.header.formato),
        ("Código", state.header.codigo_estudiante),
        ("Estudiante", f"{state.header.apellidos} {state.header.nombres}"),
        ("Carrera", f"{state.header.carrera_codigo} - {state.header.carrera_nombre}"),
        ("Modalidad / Sede", f"{state.header.modalidad} / {state.header.sede}"),
        ("Año validez", state.header.anio_validez),
        ("Procedencia", state.header.origen_nombre_procedencia),
        ("Carrera procedencia", state.header.origen_carrera_procedencia),
        ("Nota aprobatoria", state.header.origen_nota_aprobatoria),
        ("Fecha", state.header.fecha),
    ]
    for label, value in header_pairs:
        c.drawString(20 * mm, y, f"{label}: {value}")
        y -= 5 * mm

    y -= 3 * mm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, "Mapeos generados")
    y -= 5 * mm
    c.setFont("Helvetica", 8)

    if not state.mappings:
        c.drawString(20 * mm, y, "No hay mapeos registrados.")
    else:
        for idx, m in enumerate(state.mappings[:30], start=1):
            line = f"{idx}. {m.origen_codigo} -> {m.destino_codigo} | {m.destino_nombre[:60]} | score {m.score}"
            c.drawString(20 * mm, y, line)
            y -= 4.5 * mm
            if y < 20 * mm:
                c.showPage()
                y = height - 20 * mm
                c.setFont("Helvetica", 8)

    c.save()
    return path


def export_pdf(state):
    try:
        return _export_pdf_with_excel_template(state)
    except Exception:
        return _export_pdf_simple(state)
